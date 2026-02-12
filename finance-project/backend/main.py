from fastapi import FastAPI, Depends, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi import status
from sqlalchemy.orm import Session
from sqlalchemy import select, delete
from datetime import datetime, timedelta
from fastapi import File, UploadFile
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import joinedload

import database, models, schemas, crud, security
from dependencies import get_current_user
import logging
import os
import uuid
import io
import csv
from openpyxl import Workbook, load_workbook
from decimal import Decimal, InvalidOperation
from verification import send_sms, verify_sms
from config import settings

logger = logging.getLogger(__name__)

# 启动时自动在数据库建表 (C++思维：类似编译时链接)
models.Base.metadata.create_all(bind=database.engine)  # 创建所有以Base类为基类的模型类的元数据

app = FastAPI()

# 允许前端跨域访问 (必配！)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 允许所有来源，实训可以用 *
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 挂载静态文件目录（用于访问上传的头像）
os.makedirs("uploads/avatars", exist_ok=True)
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")


@app.get("/")
def read_root():
    return {"message": "Hello from FastAPI Backend!"}


# 测试 Token 认证的受保护接口
@app.get("/user/me")
def get_current_user_info(current_user: models.User = Depends(get_current_user)):
    """
    获取当前登录用户的信息（需要 Token）

    请求头需要包含：
    Authorization: Bearer <token>
    """
    return {
        "code": status.HTTP_200_OK,
        "message": "获取成功",
        "data": {
            "user_id": current_user.id,
            "username": current_user.username,
            "phone": current_user.phone or "",
            "avatar": current_user.avatar if current_user.avatar else "",
            "nickname": current_user.nickname or "",
            "email": current_user.email or "",
            "signature": current_user.signature or "",
        }
    }


# 用户登录处理函数
@app.post("/user/login")
def user_login(user: schemas.User, db: Session = Depends(database.get_db)):
    # ↑Depends依赖注入的一定是可调用函数对象，且Depends必须在处理函数中使用才会有效
    result = crud.user_login(user, db)

    # result 现在是用户对象或 None
    if result:
        # 生成 JWT Token
        access_token = security.create_access_token(
            data={"user_id": result.id, "username": result.username}
        )

        return {
            "code": status.HTTP_200_OK,
            "message": "登录成功",
            "data": {
                "user_id": result.id,
                "username": result.username,
                "phone": result.phone,
                "avatar": result.avatar if result.avatar else "",  # 头像可能为空
                "token": access_token  # 返回 JWT Token
            }
        }
    else:
        # 统一返回"用户名或密码错误"，不区分是用户不存在还是密码错误（安全考虑）
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="用户名或密码错误")
    

# 用户通过手机号登录
@app.post('/user/login_by_phone')
def login_by_phone(phone: schemas.User_phone_code, db: Session = Depends(database.get_db)):
    try:
        verify_sms(phone_number=phone.phone, verify_code=phone.verify_code)
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="验证码验证失败")
    
    result = crud.get_user_info_by_phone(phone.phone, db)
    # result 现在是用户对象或 None
    if result:
        # 生成 JWT Token
        access_token = security.create_access_token(
            data={"user_id": result.id, "username": result.username}
        )

        return {
            "code": status.HTTP_200_OK,
            "message": "登录成功",
            "data": {
                "user_id": result.id,
                "username": result.username,
                "phone": result.phone,
                "avatar": result.avatar if result.avatar else "",  # 头像可能为空
                "token": access_token  # 返回 JWT Token
            }
        }
    else:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="该手机号未注册")


# 绑定手机号发送验证码(登录注册)
@app.post('/user/send_verify_code')
def send_verify_code(phone: schemas.User_phone):
    '''
    注册/登录时phone.type=1
    更新手机时phone.type=2
    修改密码时phone.type=3
    绑定新手机时phone.type=4(这个不知道用的上，感觉和更新手机是一个意思)
    验证手机时phone.type=5
    '''
    if phone.type == 1:
        code = settings.alibabacloud_sms_template_code_rl
    elif phone.type == 2:
        code = settings.alibabacloud_sms_template_code_update
    elif phone.type == 3:
        code = settings.alibabacloud_sms_template_code_update_password
    elif phone.type == 4:
        code = settings.alibabacloud_sms_template_code_verify_new
    elif phone.type == 5:
        code = settings.alibabacloud_sms_template_code_verify_phone
    else:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='类型参数输入错误')
    try:
        send_sms(phone_number=phone.phone, code=code)
        return {
            "code": status.HTTP_200_OK,
            "message": "验证码发送成功"
        }
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="验证码发送失败")
    

# 手机号验证验证码
@app.post('/user/verify_code')
def verify_code(phone_code: schemas.User_phone_code):
    try:
        verify_sms(phone_number=phone_code.phone, verify_code=phone_code.verify_code)
        return {
            "code": status.HTTP_200_OK,
            "message": "验证码验证成功"
        }
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="验证码验证失败")


# 用户注册处理函数
@app.post("/user/register")
def user_register(user: schemas.User_register, db: Session = Depends(database.get_db)):
    result = crud.user_register(user, db)

    # result 现在是用户对象或错误代码
    # 判断是否为用户对象（成功的情况）
    if isinstance(result, models.User):
        return {
            "code": status.HTTP_200_OK,
            "message": "注册成功",
            "data": {
                "user_id": result.id,
                "username": result.username,
                "phone": result.phone,
                "avatar": result.avatar if result.avatar else ""
            }
        }
    elif result == -1:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="用户名已存在")
    elif result == -2:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="该手机号已注册")
    elif result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="注册时发生异常错误")


# 更新用户信息
@app.put("/user/update")
def update_user(user_data: schemas.User_update, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    """
    更新当前登录用户的信息（需要 Token）

    请求头需要包含：
    Authorization: Bearer <token>
    """
    result = crud.user_update(current_user.id, user_data, db)

    if isinstance(result, models.User):
        return {
            "code": status.HTTP_200_OK,
            "message": "更新成功",
            "data": {
                "user_id": result.id,
                "username": result.username,
                "phone": result.phone or "",
                "avatar": result.avatar if result.avatar else "",
                "nickname": result.nickname or "",
                "email": result.email or "",
                "signature": result.signature or ""
            }
        }
    elif result == -1:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    elif result == -2:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="该手机号已被其他用户使用")
    elif result == -3:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="该邮箱已被其他用户使用")
    elif result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="更新用户信息时发生异常错误")


# 上传头像
@app.post("/upload/avatar")
async def upload_avatar(file: UploadFile = File(...), current_user: models.User = Depends(get_current_user)):
    """
    上传用户头像（需要 Token）

    请求头需要包含：
    Authorization: Bearer <token>
    """
    # 验证文件类型
    allowed_types = ["image/jpeg", "image/png", "image/gif", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="只支持上传图片文件（jpg, png, gif, webp）")

    # 验证文件大小（限制2MB）
    file_content = await file.read()
    if len(file_content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="文件大小不能超过2MB")

    try:
        # 创建上传目录
        upload_dir = "uploads/avatars"
        os.makedirs(upload_dir, exist_ok=True)

        # 生成唯一文件名
        file_extension = os.path.splitext(file.filename)[1]
        unique_filename = f"{current_user.id}_{uuid.uuid4().hex}{file_extension}"
        file_path = os.path.join(upload_dir, unique_filename)

        # 保存文件
        with open(file_path, "wb") as f:
            f.write(file_content)

        # 返回文件访问URL（包含完整的服务器地址）
        # 在实际生产环境中，这里应该配置为真实的域名
        avatar_url = f"http://localhost:8000/uploads/avatars/{unique_filename}"

        return {
            "code": status.HTTP_200_OK,
            "message": "头像上传成功",
            "data": {
                "url": avatar_url,
                "filename": unique_filename
            }
        }
    except Exception as e:
        print(f"上传头像失败: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="上传头像时发生错误")


# 获取账单分类列表
@app.get("/category/list")
def bill_category_list(type: int, db: Session = Depends(database.get_db)):
    if type not in [1, 2]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,detail="type参数输入错误")

    result = crud.category_list(type, db)
    if result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    else:
        return {
            "code": status.HTTP_200_OK,
            "message": "成功",
            "data": result
        }


# 获取支付方式列表
@app.get("/payment_method/list")
def payment_method_list(db: Session = Depends(database.get_db)):
    result = crud.payment_method_list(db)

    if result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    else:
        return {
            "code": status.HTTP_200_OK,
            "message": "成功",
            "data": result
        }


# 创建订单
@app.post("/bill/add")
def bill_add(bill: schemas.bill_add, db: Session = Depends(database.get_db)):
    result = crud.bill_add(bill, db)

    if result == 1:
        return {
            "code": status.HTTP_200_OK,
            "message": "成功"
        }
    elif result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    elif result == -1:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该用户不存在")
    elif result == -2:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该分类不存在")
    elif result == -3:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该支付方式不存在")


# 修改订单
@app.put("/bill/update")
def bill_update(bill: schemas.bill_update, db: Session = Depends(database.get_db)):
    result = crud.bill_update(bill, db)

    if result == 1:
        return {
            "code": status.HTTP_200_OK,
            "message": "成功"
        }
    elif result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    elif result == -1:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该用户不存在")
    elif result == -2:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该分类不存在")
    elif result == -3:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该账单不存在")
    elif result == -4:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="该账单不属于此用户")
    elif result == -5:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该支付方式不存在")


# 删除账单
@app.delete("/bill/delete")
def bill_delete(bill: schemas.bill_delete, db: Session = Depends(database.get_db)):
    result = crud.bill_delete(bill, db)

    if result == 1:
        return {
            "code": status.HTTP_200_OK,
            "message": "成功"
        }
    elif result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    elif result == -1:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该用户不存在")
    elif result == -2:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该账单不存在")
    elif result == -3:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="该账单不属于此用户")


# 批量删除账单
@app.delete("/bill/batch-delete")
def bill_batch_delete(payload: schemas.bill_batch_delete, db: Session = Depends(database.get_db)):
    result = crud.bill_batch_delete(payload, db)

    # 成功：返回删除数量（兼容旧风格 code/message）
    if isinstance(result, int) and result > 0:
        return {
            "code": status.HTTP_200_OK,
            "message": "成功",
            "deleted_count": result
        }

    if result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    elif result == -1:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该用户不存在")
    elif result == -2:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该账单不存在")
    elif result == -3:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="该账单不属于此用户")
    elif result == -4:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="bill_ids不能为空")

    # 兜底
    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="未知错误")


# 获取账单列表
@app.get("/bill/list")
def bill_list(user_id: int, page:int, page_size: int, type: int, the_time: str = None, db: Session = Depends(database.get_db)):
    if page_size <15:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="页面大小太小")
    if type not in [1, 2]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="type参数输入错误")
    
    # 如果传递了 the_time，验证月份是否有效（1-12月）
    if the_time:
        try:
            year_str, month_str = the_time.split('-')
            year = int(year_str)
            month = int(month_str)

            if month < 1 or month > 12:
                raise ValueError
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="不正确的日期格式")
    
    temp = crud.get_bill_count(user_id, the_time, page_size, type, db)
    if temp == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    elif temp == -1:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该用户不存在")
    total = temp["total"]
    page_num = temp["page_num"]
    if total == 0:
        return {
            "code": status.HTTP_200_OK,
            "message": "成功",
            "total": total,
            "page_num": page_num,
            "data": []
        }

    if page < 1 or page > page_num:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="页号输入错误（为负数或是过大）")

    result = crud.bill_list(user_id, the_time, page, page_size, type, db)
    if result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    else:
        return {
            "code": status.HTTP_200_OK,
            "message": "成功",
            "total": total,
            "page_num": page_num,
            "data": result
        }


# 获取账单列表(主要用于首页页面的数据获取，将分页逻辑去除掉)
@app.get("/bill/list_first")
def bill_list_first(user_id: int, type: int, the_time: str = None, db: Session = Depends(database.get_db)):
    if type not in [1, 2]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="type参数输入错误")

    if the_time:
        try:
            year_str, month_str = the_time.split('-')
            year = int(year_str)
            month = int(month_str)

            if month < 1 or month > 12:
                raise ValueError
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="不正确的日期格式")

    result = crud.bill_list_first(user_id, the_time, type, db)
    if result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    elif result == -1:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该用户不存在")
    else:
        return {
            "code": status.HTTP_200_OK,
            "message": "成功",
            "data": result
        }

# 用于添加预算
@app.post("/budget/add")
def budget_add(budget: schemas.budget_add, db: Session = Depends(database.get_db)):
    result = crud.budget_add(budget, db)

    # 修复：检查是否返回了Budget对象（成功情况）
    if isinstance(result, models.Budget):
        return {
            "code": status.HTTP_200_OK,
            "message": "成功",
            "data": {
                "id": result.id,
                "user_id": result.user_id,
                "category_id": result.category_id,
                "is_total": result.is_total,
                "amount": float(result.amount),
                "month": result.month
            }
        }
    # 以下处理错误情况（当result为整数时）
    elif result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    elif result == -1:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="参数输入错误")
    elif result == -2:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="不正确的日期格式")
    elif result == -3:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该用户不存在")
    elif result == -4:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="总预算还未设置")
    elif result == -5:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该分类不存在")
    elif result == -6:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="一个收入分类不能设置预算")
    elif result == -7:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="当月有一个相同分类的预算已经存在")


# 用于删除预算
@app.delete("/budget/delete")
def budget_delete(budget: schemas.budget_delete, db: Session = Depends(database.get_db)):
    result = crud.budget_delete(budget, db)

    if result == 1:
        return {
            "code": status.HTTP_200_OK,
            "message": "成功"
        }
    elif result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    elif result == -1:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该用户不存在")
    elif result == -2:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该预算不存在")
    elif result == -3:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="该预算不属于此用户")


@app.put("/budget/update")
def budget_update(budget: schemas.budget_update, db: Session = Depends(database.get_db)):
    result = crud.budget_update(budget, db)

    # 修复：检查是否返回了Budget对象（成功情况）
    if isinstance(result, models.Budget):
        return {
            "code": status.HTTP_200_OK,
            "message": "成功",
            "data": {
                "id": result.id,
                "user_id": result.user_id,
                "category_id": result.category_id,
                "is_total": result.is_total,
                "amount": float(result.amount),
                "month": result.month
            }
        }
    # 以下处理错误情况（当result为整数时）
    elif result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    elif result == -1:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该用户不存在")
    elif result == -2:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该预算不存在")
    elif result == -3:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="该预算不属于此用户")


@app.get("/budget/list_month")
def budget_list_month(user_id: int, month: str, db: Session = Depends(database.get_db)):
    result = crud.budget_list_month(user_id, month, db)

    if result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    elif result == -1:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该用户不存在")
    elif result == -2:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="不正确的日期格式")
    else:
        return {
            "code": status.HTTP_200_OK,
            "message": "成功",
            "data": result
        }


"""
以下是可视化接口部分
"""
# 消费趋势分析之获取近n天数据
@app.get("/analysis/trend/days")
def get_trend_days(user_id: int = Query(..., ge=1),
                days: int = Query(..., ge=7),
                db: Session = Depends(database.get_db)
                ):
    result = crud.get_trend_days(user_id, days, db)

    if result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    elif result == -1:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该用户不存在")
    else:
        income_list, expense_list = result
        return {
            "code": status.HTTP_200_OK,
            "message": "success",
            "days": days,
            "data": {
                "income_list": income_list,
                "expense_list": expense_list
            }
        }


# 消费趋势分析之获取近n个月的数据
@app.get("/analysis/trend/months")
def get_trend_months(user_id: int = Query(..., ge=1),
                     months: int = Query(..., ge=3),
                     db: Session = Depends(database.get_db)
                     ):
    result = crud.get_trend_months(user_id, months, db)

    if result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    elif result == -1:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该用户不存在")
    else:
        income_list, expense_list = result
        return {
            "code": status.HTTP_200_OK,
            "message": "成功",
            "months": months,
            "data": {
                "income_list": income_list,
                "expense_list": expense_list
            }
        }


# 可视化界面中的“近期账单”
@app.get("/analysis/recent_bills")
def get_recent_bill(user_id: int = Query(..., ge=1),
                    db: Session = Depends(database.get_db)
                    ):
    result = crud.get_recent_bills(user_id, db)

    if result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    elif result == -1:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该用户不存在")
    else:
        return{
            "code": status.HTTP_200_OK,
            "message": "成功",
            "data": result
        }


# 消费列别占比（单个月，month=-1时为全部，month=0时为当月，month=1时为上个月，以此类推，最多往前推12个月）
@app.get("/analysis/expense_propotion_month")
def get_propotion_month(user_id: int = Query(..., ge=1),
                        month: int = Query(..., ge=-1, le=11),
                        db: Session = Depends(database.get_db)
                        ):
    result = crud.get_propotion_month(user_id, month, db)

    if result == 0:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="进行数据库业务时出错")
    elif result == -1:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail="该用户不存在")
    else:
        time, data = result
        return {
            "code": status.HTTP_200_OK,
            "message": "成功",
            "time": time,
            "data": data
        }


# 导出用户数据（CSV/Excel，需登录）
@app.get("/user/export")
def export_user_data(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(database.get_db),
):
    """
    导出用户的所有财务数据（CSV 或 Excel）
    """
    try:
        bills = db.scalars(
            select(models.Bill)
            .options(joinedload(models.Bill.bill_category), joinedload(models.Bill.payment_method))
            .where(models.Bill.user_id == current_user.id)
            .order_by(models.Bill.bill_time.desc())
        ).all()

        budgets = db.scalars(
            select(models.Budget)
            .where(models.Budget.user_id == current_user.id)
        ).all()

        rows = build_export_rows(current_user, bills, budgets)

        if format == "csv":
            csv_buffer = io.StringIO()
            writer = csv.DictWriter(csv_buffer, fieldnames=EXPORT_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)

            data = csv_buffer.getvalue().encode("utf-8")
            filename = f"finance_export_{datetime.now().strftime('%Y%m%d')}.csv"
            headers = {"Content-Disposition": f"attachment; filename={filename}"}
            return StreamingResponse(io.BytesIO(data), media_type="text/csv", headers=headers)

        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        ws.append(EXPORT_COLUMNS)
        for row in rows:
            ws.append([row.get(col, "") for col in EXPORT_COLUMNS])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"finance_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
        headers = {"Content-Disposition": f"attachment; filename={filename}"}
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"导出数据失败: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="导出数据时发生错误")


# 导入用户数据（CSV/Excel，需登录）
@app.post("/user/import")
async def import_user_data(
    file: UploadFile = File(...),
    strategy: str = Query("merge", pattern="^(merge|replace)$"),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(database.get_db),
):
    """
    导入用户的财务数据（CSV 或 Excel），格式需与导出一致
    strategy=merge 追加导入；strategy=replace 先清空再导入
    """
    filename = (file.filename or "").lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅支持 CSV 或 XLSX 文件")

    file_bytes = await file.read()
    rows: list[dict] = []

    try:
        if filename.endswith(".csv"):
            csv_text = file_bytes.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(csv_text))
            if reader.fieldnames is None:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="CSV 列头缺失")

            header_set = {h.strip() for h in reader.fieldnames if h}
            # 兼容：不同 record_type 行需要的列不同，这里只强制要求识别 record_type
            required_min = {"record_type"}
            if not required_min.issubset(header_set):
                missing = sorted(list(required_min - header_set))
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"CSV 缺少列: {', '.join(missing)}")

            rows = [row for row in reader if (row.get("record_type") or "").strip()]
        else:
            wb = load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            values = list(ws.iter_rows(values_only=True))
            if not values:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel 文件为空")

            header = [str(v).strip() if v is not None else "" for v in values[0]]
            header_set = set(header)
            required_min = {"record_type"}
            if not required_min.issubset(header_set):
                missing = sorted(list(required_min - header_set))
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Excel 缺少列: {', '.join(missing)}")

            col_index = {name: header.index(name) for name in header}
            for row_vals in values[1:]:
                record = {name: row_vals[col_index[name]] if col_index[name] < len(row_vals) else "" for name in header}
                if (record.get("record_type") or ""):
                    rows.append(record)
    except HTTPException:
        raise
    except Exception as e:
        print(f"导入解析失败: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件解析失败")

    category_ids = {c.id for c in db.scalars(select(models.Bill_Category)).all()}
    method_ids = {m.id for m in db.scalars(select(models.Payment_Method)).all()}

    bills_to_add: list[models.Bill] = []
    budgets_to_add: list[models.Budget] = []

    for idx, row in enumerate(rows, start=1):
        record_type = str(row.get("record_type") or "").strip().lower()
        if not record_type:
            continue
        try:
            if record_type == "user":
                # Excel 单元格可能是 int/float，统一转字符串再 strip
                row_user_id = parse_int(str(row.get("user_id") or "").strip(), "user_id")
                if row_user_id and row_user_id != current_user.id:
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件用户不匹配")
                continue

            if record_type == "budget":
                budget_amount = parse_decimal(str(row.get("budget_amount") or "").strip(), "budget_amount")
                budget_month = str(row.get("budget_month") or "").strip()

                # 允许空字符串/None，但如果是其它无法识别的值仍报错
                raw_is_total = row.get("budget_is_total")
                budget_is_total = parse_bool(raw_is_total)

                category_id = parse_int(str(row.get("category_id") or "").strip(), "category_id")

                if budget_amount is None or not budget_month:
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="预算数据不完整")
                if budget_is_total is None:
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="budget_is_total 格式错误")
                if not budget_is_total and category_id is None:
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="非总预算必须包含分类")
                if category_id is not None and category_id not in category_ids:
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"分类不存在: {category_id}")

                budgets_to_add.append(
                    models.Budget(
                        user_id=current_user.id,
                        category_id=category_id,
                        is_total=budget_is_total,
                        amount=float(budget_amount),
                        month=budget_month,
                    )
                )
                continue

            if record_type != "bill":
                # 未知类型直接跳过
                continue

            # record_type == bill 才进入 bill 校验
            bill_name = str(row.get("bill_name") or "").strip()
            bill_amount_raw = str(row.get("bill_amount") or "").strip()
            bill_amount = parse_decimal(bill_amount_raw, "bill_amount")
            category_id_raw = str(row.get("category_id") or "").strip()
            category_id = parse_int(category_id_raw, "category_id")
            method_id_raw = str(row.get("method_id") or "").strip()
            method_id = parse_int(method_id_raw, "method_id")
            bill_time_str = str(row.get("bill_time") or "").strip()

            # 某些情况下（例如 CSV 被 Excel/编辑器处理过）可能出现一行 record_type=bill 但其它列全空。
            # 这种行不代表有效数据：直接跳过，避免导入失败。
            if (
                not bill_name
                and not bill_amount_raw
                and not category_id_raw
                and not method_id_raw
                and not bill_time_str
                and not str(row.get("bill_remark") or "").strip()
            ):
                continue

            missing_fields = []
            if not bill_name:
                missing_fields.append("bill_name")
            if bill_amount is None:
                missing_fields.append("bill_amount")
            if category_id is None:
                missing_fields.append("category_id")
            if method_id is None:
                missing_fields.append("method_id")
            if not bill_time_str:
                missing_fields.append("bill_time")

            if missing_fields:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"账单数据不完整(缺少: {', '.join(missing_fields)})",
                )

            if category_id not in category_ids:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"分类不存在: {category_id}")
            if method_id not in method_ids:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"支付方式不存在: {method_id}")

            try:
                bill_time = datetime.fromisoformat(bill_time_str)
            except ValueError:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="bill_time 格式错误")

            bills_to_add.append(
                models.Bill(
                    user_id=current_user.id,
                    category_id=category_id,
                    method_id=method_id,
                    name=bill_name,
                    amount=float(bill_amount),
                    remark=str(row.get("bill_remark") or "").strip(),
                    bill_time=bill_time,
                )
            )
            continue
        except HTTPException as e:
            raise HTTPException(
                status_code=e.status_code,
                detail=f"第{idx}行({record_type})导入失败: {e.detail}",
            )

    if strategy == "replace":
        try:
            db.execute(delete(models.Bill).where(models.Bill.user_id == current_user.id))
            db.execute(delete(models.Budget).where(models.Budget.user_id == current_user.id))
        except Exception as e:
            db.rollback()
            print(f"清空旧数据失败: {e}")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="清空旧数据失败")

    try:
        if bills_to_add:
            db.add_all(bills_to_add)
        if budgets_to_add:
            db.add_all(budgets_to_add)
        db.commit()
        return {
            "code": status.HTTP_200_OK,
            "message": "数据导入成功",
            "data": {
                "bills_imported": len(bills_to_add),
                "budgets_imported": len(budgets_to_add),
            },
        }
    except Exception as e:
        db.rollback()
        print(f"导入数据失败: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="导入数据时发生错误")


# 清理过期数据
@app.delete("/user/clear-expired")
def clear_expired_data(payload: dict, db: Session = Depends(database.get_db)):
    """
    清理指定天数之前的账单数据
    """
    user_id = payload.get("user_id")
    days = payload.get("days", 365)

    if not user_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="user_id 不能为空")

    # 验证用户是否存在
    result = db.execute(select(models.User).where(models.User.id == user_id))
    user = result.scalar_one_or_none()

    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")

    try:
        # 计算截止日期
        cutoff_date = datetime.now() - timedelta(days=days)

        # 删除过期账单
        delete_stmt = delete(models.Bill).where(
            models.Bill.user_id == user_id,
            models.Bill.bill_time < cutoff_date
        )
        result = db.execute(delete_stmt)
        deleted_count = getattr(result, "rowcount", 0)

        db.commit()

        return {
            "code": status.HTTP_200_OK,
            "message": f"已清理 {deleted_count} 条过期账单",
            "deleted_count": deleted_count
        }
    except Exception as e:
        db.rollback()
        print(f"清理过期数据失败: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="清理数据时发生错误")


# 清空所有数据
@app.delete("/user/clear-all")
def clear_all_data(payload: dict, db: Session = Depends(database.get_db)):
    """
    清空用户的所有账单和预算数据
    """
    user_id = payload.get("user_id")

    if not user_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="user_id 不能为空")

    # 验证用户是否存在
    result = db.execute(select(models.User).where(models.User.id == user_id))
    user = result.scalar_one_or_none()

    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")

    try:
        # 删除所有账单
        bills_delete_stmt = delete(models.Bill).where(models.Bill.user_id == user_id)
        bills_result = db.execute(bills_delete_stmt)
        bills_count = getattr(bills_result, "rowcount", 0)

        # 删除所有预算
        budgets_delete_stmt = delete(models.Budget).where(models.Budget.user_id == user_id)
        budgets_result = db.execute(budgets_delete_stmt)
        budgets_count = getattr(budgets_result, "rowcount", 0)

        db.commit()

        return {
            "code": status.HTTP_200_OK,
            "message": f"已清空所有数据（{bills_count} 条账单，{budgets_count} 条预算）",
            "deleted_bills": bills_count,
            "deleted_budgets": budgets_count
        }
    except Exception as e:
        db.rollback()
        print(f"清空数据失败: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="清空数据时发生错误")


# 定义导出列
EXPORT_COLUMNS = [
    "record_type", "user_id", "username",
    "bill_id", "bill_name", "bill_amount", "bill_remark", "bill_time",
    "category_id", "category_name", "category_type",
    "method_id", "method_name",
    "budget_id", "budget_amount", "budget_month", "budget_is_total"
]


def build_export_rows(user, bills, budgets):
    """构建用于导出的数据行（保证每行都包含 EXPORT_COLUMNS 的所有列）"""

    def _blank_row():
        return {col: "" for col in EXPORT_COLUMNS}

    rows = []

    user_row = _blank_row()
    user_row.update({
        "record_type": "user",
        "user_id": user.id,
        "username": user.username,
    })
    rows.append(user_row)

    for bill in bills:
        # 关联对象可能因为懒加载/会话关闭等原因不可用，所以对 id/name 做兜底
        bill_category = getattr(bill, "bill_category", None)
        payment_method = getattr(bill, "payment_method", None)

        category_id = getattr(bill, "category_id", None) or (getattr(bill_category, "id", None) if bill_category else None)
        method_id = getattr(bill, "method_id", None) or (getattr(payment_method, "id", None) if payment_method else None)

        r = _blank_row()
        r.update({
            "record_type": "bill",
            "user_id": user.id,
            "username": user.username,
            "bill_id": bill.id,
            "bill_name": bill.name,
            "bill_amount": str(bill.amount) if bill.amount is not None else "",
            "bill_remark": bill.remark or "",
            "bill_time": bill.bill_time.isoformat() if getattr(bill, "bill_time", None) else "",
            "category_id": category_id or "",
            "category_name": (bill_category.name if bill_category else ""),
            "category_type": (bill_category.type if bill_category else ""),
            "method_id": method_id or "",
            "method_name": (payment_method.name if payment_method else ""),
        })
        rows.append(r)

    for budget in budgets:
        r = _blank_row()
        r.update({
            "record_type": "budget",
            "user_id": user.id,
            "username": user.username,
            "budget_id": budget.id,
            "budget_amount": str(budget.amount) if budget.amount is not None else "",
            "budget_month": budget.month or "",
            "budget_is_total": bool(budget.is_total),
            "category_id": budget.category_id or "",
            "category_name": budget.bill_category.name if budget.bill_category else "",
        })
        rows.append(r)

    return rows


def parse_int(value, field_name):
    """安全地将值解析为整数"""
    if value is None or value == "":
        return None
    try:
        return int(str(value).strip())
    except (ValueError, TypeError):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{field_name} 格式错误")


def parse_decimal(value, field_name):
    """安全地将值解析为 Decimal，避免 float 精度问题"""
    if value is None or value == "":
        return None
    try:
        # Excel 里可能是数字类型；统一转字符串再 Decimal
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{field_name} 格式错误")


def parse_float(value, field_name):
    """兼容旧逻辑：尽量用 Decimal 再转 float（不推荐），保留给其他接口使用"""
    dec = parse_decimal(value, field_name)
    return None if dec is None else float(dec)


def parse_bool(value):
    """安全地将值解析为布尔值"""
    if isinstance(value, bool):
        return value
    if str(value).strip().lower() in ("true", "1", "yes"):
        return True
    if str(value).strip().lower() in ("false", "0", "no", ""):
        return False
    return None
